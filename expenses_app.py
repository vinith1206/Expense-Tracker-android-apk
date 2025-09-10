from flask import Flask, request, jsonify, render_template, Response
from flask_cors import CORS
from datetime import datetime
import sqlite3
import os
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

app = Flask(__name__)
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'expenses.db')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                amount REAL NOT NULL,
                category TEXT,
                spent_at TEXT NOT NULL,
                person TEXT
            )
            """
        )
        # In case the DB already existed, ensure 'person' column exists
        cols = conn.execute("PRAGMA table_info(expenses)").fetchall()
        col_names = {c[1] for c in cols}
        if 'person' not in col_names:
            conn.execute("ALTER TABLE expenses ADD COLUMN person TEXT")

        # Budgets table: monthly budgets per category/person
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS budgets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month TEXT NOT NULL,            -- YYYY-MM
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                person TEXT
            )
            """
        )
        # Helpful index for upserts/lookups
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_budgets_unique ON budgets(month, category, IFNULL(person, ''))")
        conn.commit()
    finally:
        conn.close()

def seed_expenses_if_empty():
    conn = get_db_connection()
    try:
        row = conn.execute('SELECT COUNT(1) AS cnt FROM expenses').fetchone()
        if row and (row['cnt'] or 0) == 0:
            today = datetime.now()
            ym = today.strftime('%Y-%m')
            samples = [
                ('Rent', 18000.0, 'Rent', f'{ym}-05T09:00:00', 'Dad'),
                ('Groceries - Monthly', 6500.0, 'Groceries', f'{ym}-08T18:30:00', 'Mom'),
                ('Electricity Bill', 2200.0, 'Utilities', f'{ym}-10T10:00:00', 'Dad'),
                ('Mobile & Internet', 999.0, 'Mobile/Internet', f'{ym}-12T12:00:00', 'Dad'),
                ('Fuel - Bike', 1200.0, 'Fuel', f'{ym}-15T08:45:00', 'Dad'),
                ('Metro/Auto', 600.0, 'Transport', f'{ym}-16T19:00:00', 'Mom'),
                ('Dining Out', 750.0, 'Dining Out', f'{ym}-18T20:15:00', 'Family'),
                ('Medicines', 450.0, 'Medical', f'{ym}-20T11:20:00', 'Mom'),
                ('School Fees', 3000.0, 'Education', f'{ym}-22T09:30:00', 'Son'),
                ('Household Items', 850.0, 'Household', f'{ym}-24T17:40:00', 'Mom'),
                ('Insurance Premium', 1500.0, 'Insurance', f'{ym}-25T10:10:00', 'Dad'),
                ('SIP Investment', 2000.0, 'Savings', f'{ym}-28T07:30:00', 'Dad'),
            ]
            conn.executemany(
                'INSERT INTO expenses (title, amount, category, spent_at, person) VALUES (?, ?, ?, ?, ?)',
                samples
            )
            conn.commit()
    finally:
        conn.close()

init_db()
seed_expenses_if_empty()

@app.route('/expenses')
def expenses_page():
    return render_template('expenses.html')

@app.route('/api/expenses', methods=['GET'])
def list_expenses():
    person = (request.args.get('person') or '').strip()
    month = (request.args.get('month') or '').strip()  # YYYY-MM
    conn = get_db_connection()
    try:
        params = []
        where = []
        if person:
            where.append('person = ?')
            params.append(person)
        if month:
            start, end = _month_bounds(month)
            if start:
                where.append('datetime(spent_at) >= ? AND datetime(spent_at) < ?')
                params.extend([start, end])
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        rows = conn.execute(
            f'SELECT id, title, amount, category, spent_at, person FROM expenses{where_sql} ORDER BY datetime(spent_at) DESC, id DESC',
            params
        ).fetchall()
        return jsonify({'expenses': [dict(r) for r in rows]})
    finally:
        conn.close()

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    data = request.get_json() or {}
    title = (data.get('title') or '').strip()
    amount = data.get('amount')
    category = (data.get('category') or '').strip() or None
    spent_at = (data.get('spent_at') or datetime.now().isoformat())
    person = (data.get('person') or '').strip() or None
    if not title:
        return jsonify({'error': 'Title is required'}), 400
    try:
        amount_val = float(amount)
    except (TypeError, ValueError):
        return jsonify({'error': 'Amount must be a number'}), 400
    conn = get_db_connection()
    try:
        cur = conn.execute(
            'INSERT INTO expenses (title, amount, category, spent_at, person) VALUES (?, ?, ?, ?, ?)',
            (title, amount_val, category, spent_at, person)
        )
        conn.commit()
        new_id = cur.lastrowid
        row = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses WHERE id = ?', (new_id,)).fetchone()
        return jsonify({'expense': dict(row)}), 201
    finally:
        conn.close()

@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    conn = get_db_connection()
    try:
        cur = conn.execute('DELETE FROM expenses WHERE id = ?', (expense_id,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'Not found'}), 404
        return jsonify({'message': 'Deleted'})
    finally:
        conn.close()

@app.route('/api/expenses/<int:expense_id>', methods=['PUT', 'PATCH'])
def update_expense(expense_id):
    data = request.get_json() or {}
    fields = []
    values = []
    if 'title' in data:
        title = (data.get('title') or '').strip()
        if not title:
            return jsonify({'error': 'Title cannot be empty'}), 400
        fields.append('title = ?')
        values.append(title)
    if 'amount' in data:
        try:
            amount_val = float(data.get('amount'))
        except (TypeError, ValueError):
            return jsonify({'error': 'Amount must be a number'}), 400
        fields.append('amount = ?')
        values.append(amount_val)
    if 'category' in data:
        category = (data.get('category') or '').strip() or None
        fields.append('category = ?')
        values.append(category)
    if 'spent_at' in data:
        spent_at = (data.get('spent_at') or '').strip()
        if not spent_at:
            return jsonify({'error': 'spent_at cannot be empty'}), 400
        fields.append('spent_at = ?')
        values.append(spent_at)
    if 'person' in data:
        person = (data.get('person') or '').strip() or None
        fields.append('person = ?')
        values.append(person)
    if not fields:
        return jsonify({'error': 'No fields to update'}), 400
    conn = get_db_connection()
    try:
        values.append(expense_id)
        cur = conn.execute(f"UPDATE expenses SET {', '.join(fields)} WHERE id = ?", values)
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'Not found'}), 404
        row = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses WHERE id = ?', (expense_id,)).fetchone()
        return jsonify({'expense': dict(row)})
    finally:
        conn.close()

def _month_bounds(year_month: str):
    try:
        start = datetime.strptime(year_month + '-01', '%Y-%m-%d')
    except Exception:
        return None, None
    if start.month == 12:
        next_month = datetime(start.year + 1, 1, 1)
    else:
        next_month = datetime(start.year, start.month + 1, 1)
    return start.isoformat(), next_month.isoformat()

@app.route('/api/expenses/summary', methods=['GET'])
def summary_expenses():
    month = request.args.get('month')
    person = (request.args.get('person') or '').strip()
    conn = get_db_connection()
    try:
        if month:
            start, end = _month_bounds(month)
            if not start:
                return jsonify({'error': 'Invalid month format. Use YYYY-MM'}), 400
            if person:
                rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? AND person = ? GROUP BY category', (start, end, person)).fetchall()
                total_row = conn.execute('SELECT SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? AND person = ?', (start, end, person)).fetchone()
            else:
                rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? GROUP BY category', (start, end)).fetchall()
                total_row = conn.execute('SELECT SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ?', (start, end)).fetchone()
        else:
            if person:
                rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE person = ? GROUP BY category', (person,)).fetchall()
                total_row = conn.execute('SELECT SUM(amount) AS total FROM expenses WHERE person = ?', (person,)).fetchone()
            else:
                rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses GROUP BY category').fetchall()
                total_row = conn.execute('SELECT SUM(amount) AS total FROM expenses').fetchone()
        by_category = {(r['category'] or 'Uncategorized'): (r['total'] or 0.0) for r in rows}
        total = total_row['total'] or 0.0
        return jsonify({'total': total, 'by_category': by_category})
    finally:
        conn.close()

@app.route('/api/expenses/export.csv', methods=['GET'])
def export_expenses_csv():
    month = request.args.get('month')
    person = (request.args.get('person') or '').strip()
    conn = get_db_connection()
    try:
        if month:
            start, end = _month_bounds(month)
            if not start:
                return jsonify({'error': 'Invalid month format. Use YYYY-MM'}), 400
            if person:
                rows = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? AND person = ? ORDER BY datetime(spent_at) DESC, id DESC', (start, end, person)).fetchall()
            else:
                rows = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? ORDER BY datetime(spent_at) DESC, id DESC', (start, end)).fetchall()
        else:
            if person:
                rows = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses WHERE person = ? ORDER BY datetime(spent_at) DESC, id DESC', (person,)).fetchall()
            else:
                rows = conn.execute('SELECT id, title, amount, category, spent_at, person FROM expenses ORDER BY datetime(spent_at) DESC, id DESC').fetchall()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'title', 'amount', 'category', 'spent_at', 'person'])
        for r in rows:
            writer.writerow([r['id'], r['title'], r['amount'], r['category'] or '', r['spent_at'], r['person'] or ''])
        csv_data = output.getvalue()
        output.close()
        suffix = []
        if month:
            suffix.append(month)
        if person:
            suffix.append(person)
        suffix_str = ('-' + '-'.join(suffix)) if suffix else ''
        filename = f"expenses{suffix_str}.csv"
        return Response(csv_data, mimetype='text/csv', headers={'Content-Disposition': f'attachment; filename={filename}'})
    finally:
        conn.close()

# --- Budgets API ---
@app.route('/api/budgets', methods=['GET'])
def list_budgets():
    month = (request.args.get('month') or '').strip()
    person = (request.args.get('person') or '').strip()
    conn = get_db_connection()
    try:
        params = []
        where = []
        if month:
            where.append('month = ?')
            params.append(month)
        if person:
            where.append('IFNULL(person, "") = ?')
            params.append(person)
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        rows = conn.execute(
            f'SELECT id, month, category, amount, person FROM budgets{where_sql} ORDER BY category',
            params
        ).fetchall()
        return jsonify({'budgets': [dict(r) for r in rows]})
    finally:
        conn.close()

# --- Excel Export ---
@app.route('/api/expenses/export.xlsx', methods=['GET'])
def export_expenses_xlsx():
    month = (request.args.get('month') or '').strip()
    person = (request.args.get('person') or '').strip()
    conn = get_db_connection()
    try:
        # Fetch transactions
        params = []
        where = []
        if person:
            where.append('person = ?')
            params.append(person)
        if month:
            start, end = _month_bounds(month)
            if start:
                where.append('datetime(spent_at) >= ? AND datetime(spent_at) < ?')
                params.extend([start, end])
        where_sql = (" WHERE " + " AND ".join(where)) if where else ""
        rows = conn.execute(
            f'SELECT id, title, amount, category, spent_at, person FROM expenses{where_sql} ORDER BY datetime(spent_at) DESC, id DESC',
            params
        ).fetchall()

        # Summary by category
        if month:
            start, end = _month_bounds(month)
            if person:
                sum_rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? AND person = ? GROUP BY category', (start, end, person)).fetchall()
            else:
                sum_rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE datetime(spent_at) >= ? AND datetime(spent_at) < ? GROUP BY category', (start, end)).fetchall()
        else:
            if person:
                sum_rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses WHERE person = ? GROUP BY category', (person,)).fetchall()
            else:
                sum_rows = conn.execute('SELECT category, SUM(amount) AS total FROM expenses GROUP BY category').fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        headers = ['id', 'title', 'amount', 'category', 'spent_at', 'person']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
        for r in rows:
            ws.append([r['id'], r['title'], r['amount'], r['category'] or '', r['spent_at'], r['person'] or ''])
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(40, max(10, max_len + 2))

        ws2 = wb.create_sheet('Summary')
        ws2.append(['category', 'total'])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
        for r in sum_rows:
            ws2.append([r['category'] or 'Uncategorized', float(r['total'] or 0.0)])
        for col in ws2.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws2.column_dimensions[col[0].column_letter].width = min(30, max(10, max_len + 2))

        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()
        output.close()
        suffix = []
        if month:
            suffix.append(month)
        if person:
            suffix.append(person)
        filename = 'expenses' + ('-' + '-'.join(suffix) if suffix else '') + '.xlsx'
        return Response(data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename={filename}'})
    finally:
        conn.close()

@app.route('/api/budgets', methods=['POST'])
def upsert_budget():
    data = request.get_json() or {}
    month = (data.get('month') or '').strip()
    category = (data.get('category') or '').strip()
    person = (data.get('person') or '').strip() or None
    amount = data.get('amount')
    if not month or not category:
        return jsonify({'error': 'month and category are required'}), 400
    try:
        amount_val = float(amount)
    except (TypeError, ValueError):
        return jsonify({'error': 'amount must be a number'}), 400
    conn = get_db_connection()
    try:
        # Upsert by unique (month, category, person)
        # Normalize person to empty string for unique index
        person_key = person or ''
        row = conn.execute('SELECT id FROM budgets WHERE month = ? AND category = ? AND IFNULL(person, "") = ?', (month, category, person_key)).fetchone()
        if row:
            conn.execute('UPDATE budgets SET amount = ?, person = ? WHERE id = ?', (amount_val, person, row['id']))
        else:
            conn.execute('INSERT INTO budgets (month, category, amount, person) VALUES (?, ?, ?, ?)', (month, category, amount_val, person))
        conn.commit()
        return jsonify({'message': 'saved'})
    finally:
        conn.close()

@app.route('/api/budgets/<int:budget_id>', methods=['DELETE'])
def delete_budget(budget_id: int):
    conn = get_db_connection()
    try:
        cur = conn.execute('DELETE FROM budgets WHERE id = ?', (budget_id,))
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'Not found'}), 404
        return jsonify({'message': 'Deleted'})
    finally:
        conn.close()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5051))
    debug = os.environ.get('FLASK_ENV') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=port)

